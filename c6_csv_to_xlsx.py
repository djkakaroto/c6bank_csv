import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# Caminho para o arquivo CSV
caminho_arquivo_csv = "entrada.csv"
# Caminho para salvar o arquivo XLSX
caminho_arquivo_xlsx = "saida.xlsx"

# Leitura do arquivo CSV com ponto e vírgula como delimitador
df = pd.read_csv(caminho_arquivo_csv, delimiter=";")

# Filtrar as linhas para remover onde a coluna 'Descrição' contém 'Inclusao de Pagamento'
df = df[~df["Descrição"].str.contains("Inclusao de Pagamento", na=False)]

# Convertendo a coluna de data para o formato de data
df["Data de Compra"] = pd.to_datetime(
    df["Data de Compra"], format="%d/%m/%Y"
)  # Ajuste o formato conforme necessário

# Renomear a coluna 'Valor (em R$)' para 'Valor'
df.rename(columns={"Valor (em R$)": "Valor"}, inplace=True)

# Criar a coluna 'Descrição' unindo 'Descrição' e 'Parcela', deixando em branco se 'Parcela' for 'Única'
df["Descrição"] = df.apply(
    lambda row: (
        f"{row['Descrição']} {row['Parcela']}"
        if row["Parcela"] != "Única"
        else row["Descrição"]
    ),
    axis=1,
)
# Adicionar uma coluna 'Conta' em branco após a coluna 'Valor'
df.insert(loc=df.columns.get_loc("Valor") + 1, column="Conta", value="")

# Selecionar e reorganizar as colunas
colunas_desejadas = [
    "Data de Compra",
    "Descrição",
    "Valor",
    "Conta",
    "Categoria",
    "Nome no Cartão",
    "Final do Cartão",
    "Valor (em US$)",
    "Cotação (em R$)",
]
df_reordenado = df[colunas_desejadas]

# Salvando o DataFrame em um arquivo XLSX
df_reordenado.to_excel(caminho_arquivo_xlsx, index=False, engine="openpyxl")

# Abrindo o arquivo XLSX para aplicar formatação
wb = load_workbook(caminho_arquivo_xlsx)
ws = wb.active

# Definindo o estilo para a coluna de data
date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

# Aplicando o estilo à coluna de data
for cell in ws["A"][1:]:  # Supondo que a coluna de data está na coluna 'A'
    cell.style = date_style

# Salvando o arquivo com a formatação aplicada
wb.save(caminho_arquivo_xlsx)

print(f"Dados exportados e formatados com sucesso para {caminho_arquivo_xlsx}")
